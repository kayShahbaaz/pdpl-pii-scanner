"""
redactor.py
------------
Redaction mode for the PDPL Checker.

Given a Table and the set of cells flagged as PII, writes a redacted copy
of the source file (CSV stays CSV, XLSX stays XLSX — same sheet structure)
with flagged cells replaced by a masked placeholder.

This is meant for handing off a dataset to someone (e.g. a dev team, an
analytics vendor) without exposing the raw PII underneath.
"""

import csv
import os

from openpyxl import load_workbook

from file_readers import Table


def _mask_for_redaction(value) -> str:
    """
    Full redaction placeholder -- unlike the report's partial masking
    (which keeps a few characters for human review), a redacted output file
    replaces the value entirely so the original is not recoverable.
    """
    return "[REDACTED]"


def redact_csv(table: Table, flagged_cells: set, output_path: str):
    """
    flagged_cells: set of (row_index, column_name) tuples (0-based row index
    into table.rows) that should be redacted.
    """
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=table.columns)
        writer.writeheader()
        for i, row in enumerate(table.rows):
            new_row = dict(row)
            for col in table.columns:
                if (i, col) in flagged_cells:
                    new_row[col] = _mask_for_redaction(row.get(col))
            writer.writerow(new_row)


def redact_xlsx_sheet_inplace(wb, table: Table, flagged_cells: set):
    """
    Redacts a single sheet of an already-open openpyxl Workbook, in place.
    Row 1 is assumed to be the header (matches table.columns order).
    """
    ws = wb[table.sheet_name]
    col_index = {col: idx + 1 for idx, col in enumerate(table.columns)}  # 1-based

    for (row_i, col_name) in flagged_cells:
        excel_row = row_i + 2  # +1 for header row, +1 for 1-based indexing
        excel_col = col_index.get(col_name)
        if excel_col is not None:
            ws.cell(row=excel_row, column=excel_col).value = _mask_for_redaction(None)


def write_redacted_copy(tables_with_flags: list, source_filepath: str, output_path: str):
    """
    tables_with_flags: list of (Table, flagged_cells_set) for every sheet
    belonging to source_filepath (for CSV this list has exactly one entry).
    """
    ext = os.path.splitext(source_filepath)[1].lower()

    if ext == ".csv":
        table, flagged_cells = tables_with_flags[0]
        redact_csv(table, flagged_cells, output_path)

    elif ext == ".xlsx":
        wb = load_workbook(source_filepath)
        for table, flagged_cells in tables_with_flags:
            redact_xlsx_sheet_inplace(wb, table, flagged_cells)
        wb.save(output_path)

    else:
        raise ValueError(f"Unsupported file type for redaction: '{ext}'")
